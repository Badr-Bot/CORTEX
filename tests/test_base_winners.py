"""Tests de la base de winners — la règle d'or : les verdicts de Badr ne sont jamais écrasés."""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from agents.base_winners import upsert, est_testable, statut_vie, rafraichir, exporter_xlsx, importer_verdicts_xlsx
from agents.base_winners import testables as produits_testables  # renommé : pytest prendrait « testables » pour un test
from agents.base_winners import (notion_proprietes, notion_contenu, notion_export,
                                 importer_verdicts_notion, _page_id_depuis_url)


def test_notion_proprietes_ne_touche_jamais_aux_colonnes_de_badr():
    base = upsert({}, _pepite(), "2026-08-29", "radar quotidien")
    props = notion_proprietes(base["pestprohome.com"])
    assert props["Produit"].startswith("PestPro")
    assert props["Boutique"] == "pestprohome.com"
    assert props["✅ Testable"] == "__YES__"
    assert props["date:📅 Trouvé le:start"] == "2026-08-29"
    assert props["🇫🇷 FR"] == "🟡 PARTIEL" and props["🎚 Stade (1 libre → 5 saturé)"].startswith("2 ·")
    assert props["🔥 Statut"] == "🔥 BANGER" and props["🤖 Verdict CORTEX"] == "✅ GO TEST"
    assert not any("MON VERDICT" in k or "MES NOTES" in k for k in props)
    assert notion_export(base, "2026-08-29")[0]["icon"] == "🔥"
    # Badr, 29/08 : « éviter le vide dans les cellules » et « je vois pas quand ça dépasse deux lignes »
    textes = [v for k, v in props.items() if isinstance(v, str) and not k.startswith("date:")]
    assert all(v.strip() for v in textes), "aucune cellule ne doit être vide"
    assert all(len(v) <= 181 and "\n" not in v for v in textes), "une cellule = une ligne courte"
    # les marchés non renseignés tombent sur « à vérifier », jamais sur du vide
    nu = notion_proprietes({"boutique": "x.com", "produit": "X"})
    assert nu["🇩🇪 DE"] == "⚪ A VERIFIER" and nu["📐 TAM"] and nu["😣 Douleurs fortes"]


def test_la_fiche_page_porte_le_detail_long():
    base = upsert({}, _pepite(marches_detail={"FR": "PARTIEL — 81 pubs, Nuizoff 16 pubs depuis 10 j"},
                             pain_points=[{"douleur": "cafards malgré le ménage", "intensite": "forte",
                                           "preuve": "je n'en peux plus", "source_url": "https://reddit.com/x"}],
                             angles_non_exploites=[{"angle": "le locataire impuissant", "douleur_ciblee": "le proprio ne fait rien",
                                                    "pourquoi_personne": "tout le monde parle d'hygiène"}],
                             tam="354 pubs actives au Royaume-Uni"),
                  "2026-08-29", "radar quotidien")
    fiche = notion_contenu(base["pestprohome.com"])
    assert "## Les chiffres" in fiche and "170 pubs actives" in fiche
    assert "## Marché par marché" in fiche and "Nuizoff" in fiche
    assert "## Douleurs réelles" in fiche and "https://reddit.com/x" in fiche
    assert "## Angles que personne n'exploite" in fiche and "locataire impuissant" in fiche
    assert "## Taille du marché (TAM)" in fiche


def test_notion_export_cree_puis_ne_repousse_que_les_touches():
    base = upsert({}, _pepite(), "2026-08-29", "radar quotidien")
    assert [i["action"] for i in notion_export(base, "2026-08-29")] == ["create"]
    # une fois la page connue et rien de neuf ce jour-là : rien à pousser
    base["pestprohome.com"]["notion_page_id"] = "26ab1f9f-4c5f-80b1-8d3b-d10a6b1d2f4e"
    assert notion_export(base, "2026-08-30") == []
    base = upsert(base, _pepite(), "2026-08-30", "radar quotidien")
    assert [i["action"] for i in notion_export(base, "2026-08-30")] == ["update"]


def test_importer_verdicts_notion_relit_verdict_et_memorise_la_page():
    base = upsert({}, _pepite(), "2026-08-29", "radar quotidien")
    rows = [{"url": "https://www.notion.so/PestPro-26ab1f9f4c5f80b18d3bd10a6b1d2f4e",
             "Boutique": "pestprohome.com", "🎯 MON VERDICT": "🚫 écarté", "📝 MES NOTES": "trop vu"}]
    base, lus = importer_verdicts_notion(base, rows)
    e = base["pestprohome.com"]
    assert lus == 1 and e["verdict_badr"] == "écarté" and e["notes_badr"] == "trop vu"
    assert e["notion_page_id"] == "26ab1f9f-4c5f-80b1-8d3b-d10a6b1d2f4e"
    assert e["testable"] is False and "Badr" in e["raison_non_testable"]
    assert _page_id_depuis_url("https://app.notion.com/p/6b156b50a295410081c94286cf34321c") == "6b156b50-a295-4100-81c9-4286cf34321c"


def _pepite(**extra):
    p = {"produit": "PestPro", "boutique": "pestprohome.com", "niche": "Maison", "prix": "45 USD", "statut": "BANGER",
         "marches": {"FR": "PARTIEL", "DE": "PARTIEL"}, "stade_sophistication": 2, "verdict": "GO TEST",
         "chiffres": {"ads_actives": 170, "courbe_ads": "25 > 128", "acceleration": 5.1, "prix_eur": 41.4},
         "lien_boutique": "https://pestprohome.com"}
    p.update(extra)
    return p


def test_ajout_et_date_de_decouverte_conservee():
    base = upsert({}, _pepite(), "2026-08-29", "radar quotidien")
    base = upsert(base, _pepite(chiffres={"ads_actives": 300, "courbe_ads": "25 > 128 > 300", "acceleration": 6}), "2026-09-05", "passe hebdo")
    e = base["pestprohome.com"]
    assert e["trouve_le"] == "2026-08-29" and e["maj_le"] == "2026-09-05"
    assert [h["ads"] for h in e["historique"]] == [170, 300]
    assert e["testable"]


def test_verdict_badr_jamais_ecrase():
    base = upsert({}, _pepite(), "2026-08-29", "radar quotidien")
    base["pestprohome.com"]["verdict_badr"] = "écarté"
    base["pestprohome.com"]["notes_badr"] = "pas pour moi"
    base = upsert(base, _pepite(), "2026-09-05", "passe hebdo")
    e = base["pestprohome.com"]
    assert e["verdict_badr"] == "écarté" and e["notes_badr"] == "pas pour moi"
    assert not e["testable"] and "écarté par Badr" in e["raison_non_testable"]


def test_testable_refuse_les_marches_fermes_et_le_stade_4():
    assert est_testable({"marches": {"FR": "PRIS", "DE": "PRIS"}})[0] is False
    ok, raison = est_testable({"marches": {"FR": "LIBRE"}, "stade_sophistication": 4})
    assert not ok and "stade 4" in raison
    assert est_testable({"marches": {"FR": "LIBRE"}, "stade_sophistication": 2, "chiffres": {"prix_eur": 45}})[0]


def test_statut_vie_mort_sous_50_ads_ou_moitie_du_pic():
    assert statut_vie({"ads_actives": 30}) == "MORT"
    assert statut_vie({"ads_actives": 200, "statut": "BANGER"}, {"historique": [{"ads": 600}]}) == "MORT"
    assert statut_vie({"ads_actives": 200, "acceleration": 0.7}, {"historique": [{"ads": 220}]}) == "EN BAISSE"
    assert statut_vie({"ads_actives": 200, "acceleration": 1.5, "statut": "EXPLOSE"}, {"historique": [{"ads": 150}]}) == "EXPLOSE"


def test_refresh_sort_un_produit_mort_de_la_liste_testable():
    base = upsert({}, _pepite(), "2026-08-29", "radar quotidien")
    scan = [{"boutique": "pestprohome.com", "ads_actives": 20, "courbe_ads": "25 > 128 > 20", "acceleration": 0.2, "statut": "SOUS LE FILTRE"}]
    base = rafraichir(base, scan, "2026-09-05")
    e = base["pestprohome.com"]
    assert e["statut"] == "MORT" and not e["testable"] and len(e["historique"]) == 2
    assert produits_testables(base) == []


def test_export_et_relecture_des_verdicts(tmp_path):
    base = upsert({}, _pepite(), "2026-08-29", "radar quotidien")
    path = exporter_xlsx(base, tmp_path / "base.xlsx")
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["WINNERS"]
    entetes = [c.value for c in ws[4]]
    ws.cell(row=5, column=entetes.index("MON VERDICT") + 1, value="à tester")
    ws.cell(row=5, column=entetes.index("MES NOTES") + 1, value="tester en pack de 2")
    wb.save(path)
    base2, lus = importer_verdicts_xlsx(base, path)
    assert lus == 1
    assert base2["pestprohome.com"]["verdict_badr"] == "à tester"
    assert base2["pestprohome.com"]["notes_badr"] == "tester en pack de 2"
