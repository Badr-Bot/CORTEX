"""
CORTEX — La base de winners de Badr (leçon MASTER RESEARCH · 10 « fichier d'organisation »).

Une seule liste, tenue à jour : chaque pépite du radar quotidien y entre avec
sa date de découverte ; la passe hebdomadaire y ajoute les 10 meilleurs de la
semaine et **rafraîchit** les anciens (pubs actives, courbe, statut : toujours
banger, winner installé, en baisse, mort). Ce qui n'est plus testable sort de
la feuille WINNERS (mais reste dans HISTORIQUE).

Deux colonnes appartiennent à Badr et ne sont JAMAIS écrasées : « Mon verdict »
et « Mes notes ». Un produit qu'il a marqué « écarté » n'est plus reproposé.

Fichiers :
  data/radar/base_winners.json      la vérité (une entrée par boutique), sur main
  Notion « 🏆 BASE WINNERS — CORTEX »  la vue vivante de Badr (une ligne par boutique),
                                    tenue à jour par la routine via le connecteur Notion

Commandes :
  python -m agents.base_winners add AAAA-MM-JJ           # pépites du rapport du jour → base
  python -m agents.base_winners refresh AAAA-MM-JJ       # rafraîchit avec le scan du jour (data/radar/AAAA-MM-JJ.json)
  python -m agents.base_winners notion-export AAAA-MM-JJ # → data/radar/notion_push_AAAA-MM-JJ.json (à pousser dans Notion)
  python -m agents.base_winners import-verdicts-notion FICHIER.json  # relit MON VERDICT / MES NOTES depuis une requête Notion
  python -m agents.base_winners import-verdicts FICHIER.xlsx         # (ancien) relit depuis un Excel
  python -m agents.base_winners export [FICHIER.xlsx]                # export Excel de secours
"""

import argparse
import json
import os
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

from utils.logger import get_logger

logger = get_logger("base_winners")

# La base vit sur main. Le workflow de publication (qui tourne sur la branche
# de l'agent) pointe CORTEX_RADAR_DIR vers un checkout de main pour l'y mettre à jour.
RADAR_DIR = Path(os.getenv("CORTEX_RADAR_DIR") or (Path(__file__).parent.parent / "data" / "radar"))
BASE_PATH = RADAR_DIR / "base_winners.json"
XLSX_PATH = RADAR_DIR / "BASE-WINNERS.xlsx"

NOTION_DB_URL = "https://app.notion.com/p/6b156b50a295410081c94286cf34321c"
NOTION_DATA_SOURCE = "collection://76f47e8d-dae2-428b-843d-2f6f22305e09"

VERDICTS_BADR = ("", "à tester", "en test", "testé - winner", "écarté", "testé - mort")
STATUTS_VIE = ("BANGER", "EXPLOSE", "SANS TRAFIC", "A SURVEILLER", "STABLE", "EN BAISSE", "MORT")
PRIX_MIN_EUR = 30
MIN_ADS_VIVANT = 50


def _now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


def charger() -> dict:
    if BASE_PATH.exists():
        try:
            return json.loads(BASE_PATH.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            logger.error("base_winners.json illisible — on repart d'une base vide (l'ancien fichier est conservé en .bak)")
            BASE_PATH.replace(BASE_PATH.with_suffix(".bak.json"))
    return {}


def sauver(base: dict) -> None:
    RADAR_DIR.mkdir(parents=True, exist_ok=True)
    BASE_PATH.write_text(json.dumps(base, ensure_ascii=False, indent=1), encoding="utf-8")


# ── Règles ───────────────────────────────────────────────────────────────────

def statut_vie(chiffres: dict, precedent: dict | None = None) -> str:
    """Le produit vit-il encore ? MORT = sous le plancher de 50 ads ou courbe
    tombée à moins de la moitié de son pic (04-ecom-data-1.md:271 et :285)."""
    ads = chiffres.get("ads_actives") or 0
    if ads < MIN_ADS_VIVANT:
        return "MORT"
    pic = max([ads] + [h.get("ads", 0) for h in (precedent or {}).get("historique", [])])
    if pic and ads <= pic * 0.5:
        return "MORT"
    accel = chiffres.get("acceleration")
    if accel is not None and accel <= 0.8:
        return "EN BAISSE"
    return chiffres.get("statut") or "STABLE"


def est_testable(entree: dict) -> tuple[bool, str]:
    """Testable = ce que le protocole §1 accepte en testing produit : vivant,
    un marché de Badr ouvert, stade ≤ 3, prix ≥ 30 €, pas d'ingérable, pas
    écarté (par CORTEX ou par Badr)."""
    raisons = []
    if entree.get("verdict_badr") in ("écarté", "testé - mort"):
        raisons.append("écarté par Badr")
    if entree.get("verdict_cortex") == "ECARTER":
        raisons.append("écarté par le radar")
    if entree.get("statut") in ("MORT", "EN BAISSE"):
        raisons.append(f"courbe {entree.get('statut', '').lower()}")
    marches = entree.get("marches") or {}
    if marches and not any(v in ("LIBRE", "PARTIEL") for v in marches.values()):
        raisons.append("aucun marché ouvert")
    stade = entree.get("stade_sophistication")
    if isinstance(stade, int) and stade > 3:
        raisons.append(f"stade {stade} — nouveau mécanisme requis")
    prix = (entree.get("chiffres") or {}).get("prix_eur")
    if prix is not None and prix < PRIX_MIN_EUR:
        raisons.append(f"prix {prix} € sous l'AOV")
    if "INGERABLE" in ((entree.get("chiffres") or {}).get("alertes") or ""):
        raisons.append("ingérable")
    return (not raisons, " + ".join(raisons))


# ── Alimentation ─────────────────────────────────────────────────────────────

def upsert(base: dict, produit: dict, date: str, source: str) -> dict:
    """Ajoute ou met à jour une pépite. Les champs de Badr sont préservés."""
    dom = (produit.get("boutique") or "").lower().replace("https://", "").strip("/")
    if not dom:
        return base
    nouveau = {k: dict(v) for k, v in base.items()}
    ancien = nouveau.get(dom, {})
    chiffres = produit.get("chiffres") or ancien.get("chiffres") or {}
    if not chiffres:
        # le rapport n'a pas les chiffres (ils sont réinjectés à la publication) : on les prend dans le scan du jour
        try:
            from agents.radar_produits import chiffres_pour
            chiffres = chiffres_pour(dom, date) or {}
        except Exception:
            chiffres = {}
    entree = {
        **ancien,
        "boutique": dom,
        "produit": produit.get("produit") or ancien.get("produit", ""),
        "niche": produit.get("niche") or ancien.get("niche", ""),
        "prix": produit.get("prix") or ancien.get("prix", ""),
        "trouve_le": ancien.get("trouve_le") or date,
        "maj_le": date,
        "source": ancien.get("source") or source,
        "statut": produit.get("statut") or ancien.get("statut", ""),
        "chiffres": chiffres,
        "marches": produit.get("marches") or ancien.get("marches", {}),
        "stade_sophistication": produit.get("stade_sophistication", ancien.get("stade_sophistication")),
        "awareness": produit.get("awareness") or ancien.get("awareness", ""),
        "angle_recommande": produit.get("angle_recommande") or ancien.get("angle_recommande", ""),
        "ou_lancer": produit.get("ou_lancer") or ancien.get("ou_lancer", ""),
        "verdict_cortex": produit.get("verdict") or ancien.get("verdict_cortex", ""),
        "verdict_pourquoi": produit.get("verdict_pourquoi") or ancien.get("verdict_pourquoi", ""),
        "tam": produit.get("tam") or ancien.get("tam", ""),
        "recurrent": produit.get("recurrent") or ancien.get("recurrent", ""),
        "difficulte": produit.get("difficulte") or ancien.get("difficulte", ""),
        "ca_jour_estime": produit.get("ca_jour_estime") or ancien.get("ca_jour_estime", ""),
        "budget_test": produit.get("budget_test") or ancien.get("budget_test", ""),
        "angle_concurrent": produit.get("angle_concurrent") or ancien.get("angle_concurrent", ""),
        "pain_points": produit.get("pain_points") or ancien.get("pain_points", []),
        "angles_non_exploites": produit.get("angles_non_exploites") or ancien.get("angles_non_exploites", []),
        "marches_detail": produit.get("marches_detail") or ancien.get("marches_detail", {}),
        "lien_boutique": produit.get("lien_boutique") or ancien.get("lien_boutique") or f"https://{dom}",
        "lien_adlibrary": produit.get("lien_adlibrary") or ancien.get("lien_adlibrary", ""),
        "verdict_badr": ancien.get("verdict_badr", ""),
        "notes_badr": ancien.get("notes_badr", ""),
        "historique": ancien.get("historique", []),
    }
    if chiffres:
        point = {"date": date, "ads": chiffres.get("ads_actives"), "courbe": chiffres.get("courbe_ads"),
                 "statut": entree["statut"]}
        if not entree["historique"] or entree["historique"][-1].get("date") != date:
            entree["historique"] = entree["historique"] + [point]
    ok, raison = est_testable(entree)
    entree["testable"] = ok
    entree["raison_non_testable"] = raison
    nouveau[dom] = entree
    return nouveau


def ajouter_depuis_rapport(base: dict, rapport: dict, date: str) -> dict:
    for p in (rapport.get("ecommerce", {}).get("radar_produits") or []):
        base = upsert(base, p, date, "radar quotidien")
    return base


def rafraichir(base: dict, scan: list[dict], date: str) -> dict:
    """Met à jour les chiffres des produits de la base présents dans un scan."""
    par_dom = {p["boutique"].lower(): p for p in scan}
    nouveau = {k: dict(v) for k, v in base.items()}
    for dom, entree in nouveau.items():
        p = par_dom.get(dom)
        if not p:
            continue
        chiffres = {k: p.get(k) for k in ("ads_actives", "courbe_ads", "acceleration", "delta_ads_7j", "age_jours",
                                           "semaines_diffusion", "visites_mois", "pays_pub", "nb_skus", "prix",
                                           "prix_eur", "alertes", "statut") if k in p}
        statut = statut_vie(chiffres, entree)
        entree.update({"chiffres": {**entree.get("chiffres", {}), **chiffres}, "statut": statut, "maj_le": date})
        entree["historique"] = entree.get("historique", []) + [
            {"date": date, "ads": chiffres.get("ads_actives"), "courbe": chiffres.get("courbe_ads"), "statut": statut}]
        ok, raison = est_testable(entree)
        entree["testable"], entree["raison_non_testable"] = ok, raison
    return nouveau


def testables(base: dict) -> list[dict]:
    rows = [e for e in base.values() if e.get("testable")]
    rows.sort(key=lambda e: (e.get("verdict_cortex") != "GO TEST", -(e.get("chiffres", {}).get("ads_actives") or 0)))
    return rows


# ── Notion ───────────────────────────────────────────────────────────────────
# La routine n'a pas de réseau mais a le connecteur Notion : ce module prépare
# les propriétés (notion-export) et relit les verdicts (import-verdicts-notion) ;
# c'est l'agent qui appelle create-pages / update-page avec ce JSON.

# Les options Notion portent un emoji (Badr veut repérer d'un coup d'œil) :
# ici la valeur brute → l'option Notion. À l'import on fait l'inverse en
# retirant tout ce qui précède la première lettre.
STATUT_NOTION = {"BANGER": "🔥 BANGER", "EXPLOSE": "🚀 EXPLOSE", "SANS TRAFIC": "👻 SANS TRAFIC",
                 "A SURVEILLER": "👀 A SURVEILLER", "STABLE": "➖ STABLE", "EN BAISSE": "📉 EN BAISSE", "MORT": "💀 MORT"}
VERDICT_NOTION = {"GO TEST": "✅ GO TEST", "A SURVEILLER": "👀 A SURVEILLER", "ECARTER": "❌ ECARTER"}
MARCHE_NOTION = {"LIBRE": "🟢 LIBRE", "PARTIEL": "🟡 PARTIEL", "PRIS": "🔴 PRIS", "A VERIFIER": "⚪ A VERIFIER"}
VERDICT_BADR_NOTION = {"à tester": "🧪 à tester", "en test": "⏳ en test", "testé - winner": "🏆 testé - winner",
                       "testé - mort": "💀 testé - mort", "écarté": "🚫 écarté"}
ICONE_STATUT = {"BANGER": "🔥", "EXPLOSE": "🚀", "SANS TRAFIC": "👻", "A SURVEILLER": "👀", "STABLE": "➖",
                "EN BAISSE": "📉", "MORT": "💀"}
DIFFICULTE_NOTION = {"facile": "🟢 facile", "moyen": "🟡 moyen", "dur": "🔴 dur"}
STADE_NOTION = {1: "1 · personne", 2: "2 · quelques récents — OK", 3: "3 · tout le monde dit pareil",
                4: "4 · saturé", 5: "5 · nouveau mécanisme requis"}
AWARENESS_NOTION = {"inconnu ici": "inconnu ici → éduquer", "déjà connu ici": "déjà connu → nouvel angle"}
RECURRENT_NOTION = {"oui": "oui · consommable", "non": "non · achat unique"}
COL_VERDICT_BADR, COL_NOTES_BADR = "🎯 MON VERDICT", "📝 MES NOTES"
VIDE = "—"          # Badr : « éviter le vide dans les cellules »
A_VERIFIER = "à vérifier"
MAX_CELLULE = 180   # au-delà, le texte vit dans la page, pas dans la colonne


def sans_emoji(valeur) -> str:
    """'🔥 BANGER' → 'BANGER', '🚫 écarté' → 'écarté'. Retire tout ce qui précède la première lettre."""
    s = str(valeur or "").strip()
    i = 0
    while i < len(s) and not s[i].isalpha():
        i += 1
    return s[i:].strip()


def _difficulte(entree: dict) -> str | None:
    d = (entree.get("difficulte") or "").lower()
    for cle, val in DIFFICULTE_NOTION.items():
        if cle in d:
            return val
    return None


def notion_icone(entree: dict) -> str:
    if entree.get("verdict_badr") == "testé - winner":
        return "🏆"
    return ICONE_STATUT.get(entree.get("statut") or "", "🛍️")


def _court(texte, defaut: str = VIDE) -> str:
    """Une cellule = une ligne lisible. Le texte long vit dans la page."""
    t = " ".join(str(texte or "").split())
    if not t:
        return defaut
    return t if len(t) <= MAX_CELLULE else t[:MAX_CELLULE - 1].rstrip(" ,;.") + "…"


def _premiere_pub(entree: dict) -> str | None:
    """Date de la première pub = aujourd'hui - semaines de diffusion (ce que TrendTrack sait)."""
    c = entree.get("chiffres") or {}
    semaines = c.get("semaines_diffusion")
    base_date = entree.get("maj_le") or entree.get("trouve_le")
    if not semaines or not base_date:
        return None
    try:
        d = datetime.strptime(base_date, "%Y-%m-%d")
    except ValueError:
        return None
    return (d - timedelta(weeks=int(semaines))).strftime("%Y-%m-%d")


def _marches_detail(entree: dict) -> str:
    """« Pourquoi DE partiel alors que tu dis qu'il n'y a personne » : la preuve, marché par marché."""
    detail = entree.get("marches_detail") or {}
    m = entree.get("marches") or {}
    if detail:
        return _court(" · ".join(f"{pays} : {txt}" for pays, txt in detail.items() if txt))
    if entree.get("marche_fr_detail"):
        return _court(entree["marche_fr_detail"])
    if m:
        return _court(" · ".join(f"{p} {v}" for p, v in m.items()) + " — détail non enregistré, relance le contrôle Meta Ad Library")
    return "contrôle pays pas encore fait"


def _resume_angles(entree: dict) -> str:
    angles = entree.get("angles_non_exploites") or []
    if not angles:
        return "aucun angle libre identifié — l'enquête forums n'a rien donné"
    return _court(" · ".join(a.get("angle", "") for a in angles if a.get("angle")))


def _resume_douleurs(entree: dict) -> str:
    douleurs = [d for d in (entree.get("pain_points") or []) if d.get("intensite") == "forte"] or (entree.get("pain_points") or [])
    if not douleurs:
        return "pas de douleur forte trouvée dans les forums"
    return _court(" · ".join(d.get("douleur", "") for d in douleurs))


def _score_sur_9(entree: dict) -> int | None:
    ok = entree.get("criteres_ok")
    return len(ok) if isinstance(ok, list) and ok else None


def notion_proprietes(entree: dict) -> dict:
    """Propriétés Notion d'une entrée (jamais MON VERDICT / MES NOTES : ils sont à Badr).
    Règle Badr du 29/08 : aucune cellule vide, une seule ligne par cellule."""
    c = entree.get("chiffres") or {}
    m = entree.get("marches") or {}
    props = {
        "Produit": entree.get("produit") or entree.get("boutique", ""),
        "Boutique": entree.get("boutique", ""),
        "🤖 Verdict CORTEX": VERDICT_NOTION.get(entree.get("verdict_cortex") or ""),
        "🔥 Statut": STATUT_NOTION.get(entree.get("statut") or ""),
        "✅ Testable": "__YES__" if entree.get("testable") else "__NO__",
        "date:📅 Trouvé le:start": entree.get("trouve_le"), "date:📅 Trouvé le:is_datetime": 0,
        "date:🔄 Mis à jour:start": entree.get("maj_le"), "date:🔄 Mis à jour:is_datetime": 0,
        "🧩 Niche": _court(entree.get("niche")),
        "💵 Prix vu": _court(entree.get("prix")),
        "💶 Prix EUR": c.get("prix_eur"),
        "📣 Pubs actives": c.get("ads_actives"),
        "📈 Accélération x4 sem": c.get("acceleration"),
        "📉 Courbe 5 sem": _court(c.get("courbe_ads"), "courbe non disponible"),
        "⏳ Âge (jours)": c.get("age_jours"),
        "👣 Visites / mois": c.get("visites_mois"),
        "🌐 Pays de leurs pubs": _court(c.get("pays_pub"), "pays non indexés"),
        "🇫🇷 FR": MARCHE_NOTION.get(m.get("FR") or "", MARCHE_NOTION["A VERIFIER"]),
        "🇩🇪 DE": MARCHE_NOTION.get(m.get("DE") or "", MARCHE_NOTION["A VERIFIER"]),
        "🇪🇸 ES": MARCHE_NOTION.get(m.get("ES") or "", MARCHE_NOTION["A VERIFIER"]),
        "🇬🇧 GB": MARCHE_NOTION.get(m.get("GB") or "", MARCHE_NOTION["A VERIFIER"]),
        "🌍 Marchés (détail)": _marches_detail(entree),
        "🎚 Stade (1 libre → 5 saturé)": STADE_NOTION.get(entree.get("stade_sophistication")),
        "🧠 Awareness": AWARENESS_NOTION.get(entree.get("awareness") or "", A_VERIFIER),
        "📐 TAM": _court(entree.get("tam"), "TAM pas encore mesuré"),
        "🔁 Récurrent": RECURRENT_NOTION.get(str(entree.get("recurrent") or "").lower()[:3].rstrip(), A_VERIFIER),
        "🌍 Où lancer": _court(entree.get("ou_lancer")),
        "💰 CA / jour estimé": _court(entree.get("ca_jour_estime"), "non estimable"),
        "⚡ Difficulté": _difficulte(entree),
        "💡 Angle recommandé": _court(entree.get("angle_recommande")),
        "🗣 Angle du concurrent": _court(entree.get("angle_concurrent"), "pubs concurrentes pas encore lues"),
        "💥 Angles non exploités": _resume_angles(entree),
        "😣 Douleurs fortes": _resume_douleurs(entree),
        "🧾 Budget test": _court(entree.get("budget_test"), "CBO 100-300 €/j, décision à 48 h"),
        "⭐ Score /9": _score_sur_9(entree),
        "🔗 Lien boutique": entree.get("lien_boutique") or None,
        "📺 Leurs pubs": entree.get("lien_adlibrary") or None,
    }
    premiere = _premiere_pub(entree)
    if premiere:
        props["date:🎬 Première pub:start"] = premiere
        props["date:🎬 Première pub:is_datetime"] = 0
    return {k: v for k, v in props.items() if v is not None}


def notion_contenu(entree: dict) -> str:
    """La fiche complète (corps de la page) : tout ce qui ne tient pas sur une ligne."""
    c = entree.get("chiffres") or {}
    L: list[str] = []

    def bloc(titre: str, corps: str):
        if corps:
            L.extend([f"## {titre}", "", corps, ""])

    chiffres = [f"**{c.get('ads_actives', '?')} pubs actives**"]
    if c.get("acceleration") is not None:
        chiffres.append(f"×{c['acceleration']} en 4 semaines")
    if c.get("courbe_ads"):
        chiffres.append(f"courbe {c['courbe_ads']}")
    if c.get("age_jours") is not None:
        chiffres.append(f"boutique de {c['age_jours']} jours")
    if c.get("semaines_diffusion"):
        chiffres.append(f"{c['semaines_diffusion']} semaines de diffusion")
    if c.get("visites_mois") is not None:
        chiffres.append(f"{c['visites_mois']} visites/mois")
    if c.get("nb_skus") is not None:
        chiffres.append(f"{c['nb_skus']} produits au catalogue")
    bloc("Les chiffres", " · ".join(chiffres))

    bloc(f"Pourquoi CORTEX dit {entree.get('verdict_cortex') or '…'}", entree.get("verdict_pourquoi", ""))
    bloc("Où lancer", entree.get("ou_lancer", ""))

    detail = entree.get("marches_detail") or {}
    if detail:
        bloc("Marché par marché", "\n".join(f"- **{p}** — {t}" for p, t in detail.items() if t))
    elif entree.get("marche_fr_detail"):
        bloc("Marché par marché", entree["marche_fr_detail"])

    bloc("Combien ça peut faire", entree.get("ca_jour_estime", ""))
    bloc("Taille du marché (TAM)", entree.get("tam", ""))
    bloc("Angle recommandé", entree.get("angle_recommande", ""))
    bloc("Ce que disent les concurrents", entree.get("angle_concurrent", ""))

    douleurs = entree.get("pain_points") or []
    if douleurs:
        bloc("Douleurs réelles (scraping forums)", "\n".join(
            f"- **{d.get('intensite', '?')}** — {d.get('douleur', '')}" +
            (f"\n  > {d.get('preuve')}" if d.get("preuve") else "") +
            (f"\n  [source]({d['source_url']})" if d.get("source_url") else "")
            for d in douleurs))

    angles = entree.get("angles_non_exploites") or []
    if angles:
        bloc("Angles que personne n'exploite", "\n".join(
            f"- **{a.get('angle', '')}**" +
            (f"\n  Douleur ciblée : {a['douleur_ciblee']}" if a.get("douleur_ciblee") else "") +
            (f"\n  Pourquoi personne ne le fait : {a['pourquoi_personne']}" if a.get("pourquoi_personne") else "")
            for a in angles))

    bloc("Budget de test", entree.get("budget_test", ""))
    if entree.get("raison_non_testable"):
        bloc("Pourquoi il n'est plus testable", entree["raison_non_testable"])

    hist = entree.get("historique") or []
    if hist:
        bloc("Historique", "\n".join(
            f"- {h.get('date')} — {h.get('ads') if h.get('ads') is not None else '?'} pubs · {h.get('courbe') or '?'} · {h.get('statut') or ''}"
            for h in hist))
    return "\n".join(L).strip()


def notion_export(base: dict, date: str) -> list[dict]:
    """Ce qu'il faut pousser dans Notion : les entrées touchées ce jour ou jamais créées."""
    out = []
    for dom, e in sorted(base.items()):
        if e.get("maj_le") != date and e.get("notion_page_id"):
            continue
        out.append({
            "boutique": dom,
            "notion_page_id": e.get("notion_page_id", ""),
            "action": "update" if e.get("notion_page_id") else "create",
            "icon": notion_icone(e),
            "properties": notion_proprietes(e),
            "content": notion_contenu(e),
        })
    return out


def _page_id_depuis_url(url: str) -> str:
    """https://www.notion.so/xxx-26ab1f9f4c5f80b18d3bd10a6b1d2f4e → 26ab1f9f-4c5f-80b1-8d3b-d10a6b1d2f4e"""
    brut = (url or "").rstrip("/").split("/")[-1].split("?")[0].split("-")[-1].replace("-", "")
    if len(brut) != 32:
        return ""
    return f"{brut[:8]}-{brut[8:12]}-{brut[12:16]}-{brut[16:20]}-{brut[20:]}"


def importer_verdicts_notion(base: dict, lignes: list[dict]) -> tuple[dict, int]:
    """Relit MON VERDICT / MES NOTES depuis les lignes d'une requête Notion
    (SELECT url, Boutique, "MON VERDICT", "MES NOTES" …). Mémorise aussi l'id
    de page pour les prochaines mises à jour."""
    nouveau = {k: dict(v) for k, v in base.items()}
    lus = 0

    def _col(row: dict, fin: str):
        """Tolère les noms avec ou sans emoji ('🎯 MON VERDICT' ou 'MON VERDICT')."""
        for k, v in row.items():
            if sans_emoji(k).upper() == fin:
                return v
        return None

    for row in lignes:
        dom = str(row.get("Boutique") or _col(row, "BOUTIQUE") or "").strip().lower()
        if dom not in nouveau:
            continue
        page_id = _page_id_depuis_url(row.get("url") or row.get("notion_page_id") or "")
        if page_id:
            nouveau[dom]["notion_page_id"] = page_id
        verdict = sans_emoji(_col(row, "MON VERDICT")).lower()
        notes = str(_col(row, "MES NOTES") or "").strip()
        if verdict != nouveau[dom].get("verdict_badr", "") or notes != nouveau[dom].get("notes_badr", ""):
            nouveau[dom]["verdict_badr"] = verdict
            nouveau[dom]["notes_badr"] = notes
            ok, raison = est_testable(nouveau[dom])
            nouveau[dom]["testable"], nouveau[dom]["raison_non_testable"] = ok, raison
            lus += 1
    return nouveau, lus


# ── Excel ────────────────────────────────────────────────────────────────────

COLONNES = [
    ("trouve_le", "TROUVÉ LE", 11), ("maj_le", "MAJ LE", 11), ("produit", "PRODUIT", 42), ("boutique", "BOUTIQUE", 24),
    ("niche", "NICHE", 16), ("prix", "PRIX VU", 12), ("statut", "STATUT", 13), ("_ads", "PUBS ACTIVES", 10),
    ("_courbe", "COURBE 5 SEM.", 26), ("_accel", "×4 SEM.", 8), ("_age", "ÂGE (j)", 8), ("_pays", "PAYS DE LEURS PUBS", 26),
    ("_marches", "MARCHÉS FR/DE/ES/GB", 30), ("stade_sophistication", "STADE", 7), ("awareness", "AWARENESS", 14),
    ("verdict_cortex", "VERDICT CORTEX", 14), ("ou_lancer", "OÙ LANCER", 40), ("angle_recommande", "ANGLE", 50),
    ("_testable", "TESTABLE ?", 12), ("raison_non_testable", "SINON POURQUOI", 30),
    ("verdict_badr", "MON VERDICT", 14), ("notes_badr", "MES NOTES", 40),
    ("lien_boutique", "BOUTIQUE (lien)", 14), ("lien_adlibrary", "AD LIBRARY", 14),
]


def _valeur(entree: dict, cle: str):
    c = entree.get("chiffres") or {}
    if cle == "_ads": return c.get("ads_actives")
    if cle == "_courbe": return c.get("courbe_ads")
    if cle == "_accel": return c.get("acceleration")
    if cle == "_age": return c.get("age_jours")
    if cle == "_pays": return c.get("pays_pub")
    if cle == "_marches": return " · ".join(f"{k} {v}" for k, v in (entree.get("marches") or {}).items())
    if cle == "_testable": return "OUI" if entree.get("testable") else "non"
    return entree.get(cle, "")


def exporter_xlsx(base: dict, path: Path = XLSX_PATH) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    fills = {"BANGER": "FFFDECEA", "EXPLOSE": "FFEAF5EB", "SANS TRAFIC": "FFF3E5F5", "A SURVEILLER": "FFE9F2FC",
             "STABLE": "FFEDF0F2", "EN BAISSE": "FFFBE9E7", "MORT": "FFEEEEEE"}

    def feuille(nom: str, lignes: list[dict], sous_titre: str):
        ws = wb.create_sheet(nom)
        ws["A1"] = nom
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = sous_titre
        ws["A2"].font = Font(italic=True, size=9, color="FF777777")
        for i, (_, titre, larg) in enumerate(COLONNES, 1):
            c = ws.cell(row=4, column=i, value=titre)
            c.font = Font(bold=True, color="FFFFFFFF", size=9)
            c.fill = PatternFill("solid", fgColor="FF2B2B2B")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = larg
        for r, e in enumerate(lignes, 5):
            teinte = fills.get(e.get("statut", ""), "FFFFFFFF")
            for i, (cle, _, _) in enumerate(COLONNES, 1):
                v = _valeur(e, cle)
                c = ws.cell(row=r, column=i, value=v if v not in ("", None) else None)
                c.font = Font(size=9)
                c.fill = PatternFill("solid", fgColor=teinte)
                c.alignment = Alignment(vertical="center", wrap_text=cle in ("angle_recommande", "ou_lancer", "notes_badr", "raison_non_testable"))
                if cle in ("lien_boutique", "lien_adlibrary") and v:
                    c.value = "ouvrir >"
                    c.hyperlink = v
                    c.font = Font(color="FF0563C1", underline="single", size=9)
                if cle in ("verdict_badr", "notes_badr"):
                    c.fill = PatternFill("solid", fgColor="FFFFF8E1")
        ws.freeze_panes = "D5"
        if lignes:
            ws.auto_filter.ref = f"A4:{get_column_letter(len(COLONNES))}{4 + len(lignes)}"

    feuille("WINNERS", testables(base),
            "Uniquement les produits encore testables (vivants, un marché ouvert, stade ≤ 3, prix ≥ 30 €). "
            "MON VERDICT et MES NOTES sont à toi : jamais écrasés. Verdicts possibles : à tester · en test · winner · écarté · testé - mort.")
    tous = sorted(base.values(), key=lambda e: e.get("trouve_le", ""), reverse=True)
    feuille("HISTORIQUE", tous, "Tout ce que le radar a proposé, testable ou non, avec la raison.")
    leg = wb.create_sheet("LEGENDE")
    for i, ligne in enumerate([
        "STATUT : BANGER = trafic 0 + courbe ×2 · EXPLOSE = courbe ×2, trafic visible · SANS TRAFIC = fenêtre ouverte, courbe plate · "
        "A SURVEILLER = 50-149 pubs · STABLE = ≥150 pubs, courbe plate · EN BAISSE = courbe ≤ ×0,8 · MORT = < 50 pubs ou courbe tombée à la moitié du pic",
        "STADE (leçon 33) : 1 = personne · 2 = 1-4 concurrents récents (OK) · 3 = tout le monde dit pareil · 4-5 = saturé",
        "MARCHÉS : LIBRE = personne · PARTIEL = 1-4 récents ou mal exécutés (OK pour lancer) · PRIS = ≥5 ou un acteur dominant · A VERIFIER",
        "TESTABLE : vivant + un marché ouvert + stade ≤ 3 + prix ≥ 30 € + pas d'ingérable + pas écarté (par le radar ou par toi)",
        "Le radar ne repropose jamais un produit marqué « écarté » ou « testé - mort » dans MON VERDICT.",
    ], 1):
        leg.cell(row=i, column=1, value=ligne).alignment = Alignment(wrap_text=True)
    leg.column_dimensions["A"].width = 140
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def importer_verdicts_xlsx(base: dict, path: Path) -> tuple[dict, int]:
    """Relit MON VERDICT / MES NOTES (par nom d'en-tête) dans les deux feuilles."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    nouveau = {k: dict(v) for k, v in base.items()}
    lus = 0
    for nom in ("WINNERS", "HISTORIQUE"):
        if nom not in wb.sheetnames:
            continue
        ws = wb[nom]
        rows = ws.iter_rows(min_row=4, values_only=True)
        entetes = [str(c).strip() if c else "" for c in next(rows, [])]
        if "BOUTIQUE" not in entetes or "MON VERDICT" not in entetes:
            continue
        ib, iv = entetes.index("BOUTIQUE"), entetes.index("MON VERDICT")
        inote = entetes.index("MES NOTES") if "MES NOTES" in entetes else None
        for row in rows:
            dom = str(row[ib] or "").strip().lower()
            if dom not in nouveau:
                continue
            verdict = str(row[iv] or "").strip().lower()
            notes = str(row[inote] or "").strip() if inote is not None else nouveau[dom].get("notes_badr", "")
            if verdict or notes:
                nouveau[dom]["verdict_badr"] = verdict
                nouveau[dom]["notes_badr"] = notes
                ok, raison = est_testable(nouveau[dom])
                nouveau[dom]["testable"], nouveau[dom]["raison_non_testable"] = ok, raison
                lus += 1
    return nouveau, lus


# ── CLI ──────────────────────────────────────────────────────────────────────

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = parser.add_subparsers(dest="cmd", required=True)
    p = sub.add_parser("add"); p.add_argument("date")
    p = sub.add_parser("refresh"); p.add_argument("date")
    p = sub.add_parser("import-verdicts"); p.add_argument("fichier")
    p = sub.add_parser("import-verdicts-notion"); p.add_argument("fichier")
    p = sub.add_parser("notion-export"); p.add_argument("date")
    p = sub.add_parser("export"); p.add_argument("fichier", nargs="?", default=str(XLSX_PATH))
    args = parser.parse_args(argv)

    base = charger()
    if args.cmd == "notion-export":
        items = notion_export(base, args.date)
        out = RADAR_DIR / f"notion_push_{args.date}.json"
        out.write_text(json.dumps(items, ensure_ascii=False, indent=1), encoding="utf-8")
        logger.info(f"{len(items)} page(s) à pousser dans Notion → {out.name} "
                    f"({sum(i['action'] == 'create' for i in items)} à créer)")
        return 0
    if args.cmd == "export":
        path = exporter_xlsx(base, Path(args.fichier))
        logger.info(f"Export Excel : {path}")
        return 0
    if args.cmd == "add":
        rapport_path = Path(os.getenv("CORTEX_RAPPORT") or (Path(__file__).parent.parent / "data" / "cowork" / f"rapport_{args.date}.json"))
        if not rapport_path.exists():
            logger.error(f"{rapport_path} introuvable"); return 1
        base = ajouter_depuis_rapport(base, json.loads(rapport_path.read_text(encoding="utf-8")), args.date)
    elif args.cmd == "refresh":
        scan_path = RADAR_DIR / f"{args.date}.json"
        if not scan_path.exists():
            logger.error(f"{scan_path} introuvable"); return 1
        base = rafraichir(base, json.loads(scan_path.read_text(encoding="utf-8")), args.date)
    elif args.cmd == "import-verdicts":
        base, lus = importer_verdicts_xlsx(base, Path(args.fichier))
        logger.info(f"{lus} verdict(s) de Badr relus")
    elif args.cmd == "import-verdicts-notion":
        brut = json.loads(Path(args.fichier).read_text(encoding="utf-8"))
        lignes = brut.get("rows") or brut.get("results") or brut if isinstance(brut, dict) else brut
        base, lus = importer_verdicts_notion(base, lignes if isinstance(lignes, list) else [])
        logger.info(f"{lus} verdict(s) de Badr relus depuis Notion")
    sauver(base)
    logger.info(f"Base : {len(base)} produits, {len(testables(base))} testables")
    return 0


if __name__ == "__main__":
    sys.exit(main())
